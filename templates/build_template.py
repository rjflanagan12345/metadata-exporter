from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.comments import Comment

wb = Workbook()

# ─── COLOR PALETTE ──────────────────────────────────────────────────────────
NAVY       = "1F3864"
DARK_BLUE  = "2E5FA3"
MED_BLUE   = "4472C4"
LIGHT_BLUE = "D6E4F7"
PALE_BLUE  = "EEF4FC"
REQUIRED   = "C00000"   # dark red
OPT_GRN    = "375623"   # dark green
YELLOW     = "FFF2CC"
WHITE      = "FFFFFF"
GREY_HDR   = "F2F2F2"
GREY_LINE  = "D9D9D9"
SECTION_BG = "BDD7EE"

def hdr_font(color="FFFFFF", size=9, bold=True):
    return Font(name="Arial", bold=bold, color=color, size=size)

def cell_font(size=9, bold=False, color="000000"):
    return Font(name="Arial", size=size, bold=bold, color=color)

def fill(hex_color):
    return PatternFill("solid", start_color=hex_color, fgColor=hex_color)

def thin_border():
    s = Side(style="thin", color=GREY_LINE)
    return Border(left=s, right=s, top=s, bottom=s)

def wrap_align(horizontal="left", vertical="top"):
    return Alignment(horizontal=horizontal, vertical=vertical, wrap_text=True)

# ════════════════════════════════════════════════════════════════════════════
# SHEET 1 — METADATA ENTRY
# ════════════════════════════════════════════════════════════════════════════
ws = wb.active
ws.title = "Metadata"
ws.sheet_view.showGridLines = True
ws.freeze_panes = "A4"   # freeze first 3 header rows

# ── Column definitions ────────────────────────────────────────────────────
# (header, section, required, format_note, example, width)
COLUMNS = [
    # ── IDENTIFICATION ──────────────────────────────────────────────────
    ("ISBN10",              "IDENTIFICATION", False,
     "10-digit ISBN. Include leading zeros. Leave blank if none.",
     "099949600X", 14),
    ("ISBN13",              "IDENTIFICATION", True,
     "13-digit ISBN starting with 978 or 979. No hyphens or spaces. PRIMARY key used to match cover images and sync to CORE.",
     "9780999496008", 18),
    ("GTIN",                "IDENTIFICATION", False,
     "Same as ISBN13 in most cases. Use if different.",
     "9780999496008", 16),
    ("UPC",                 "IDENTIFICATION", False,
     "Universal Product Code. Only if product has a separate UPC barcode.",
     "", 14),
    # Work ID removed — system auto-populates from ISBN13
    ("eISBN10 for Print",   "IDENTIFICATION", False,
     "eBook ISBN10 linked to this print edition.",
     "", 16),
    ("eISBN13 for Print",   "IDENTIFICATION", False,
     "eBook ISBN13 linked to this print edition.",
     "", 18),

    # ── TITLE ────────────────────────────────────────────────────────────
    ("Title",               "TITLE", True,
     "Full title exactly as it appears on cover. Do not include subtitle here.",
     "The Best Book", 30),
    ("Subtitle",            "TITLE", False,
     "Subtitle only. Do not repeat the main title.",
     "A Story of Friendship", 30),
    ("Series Name",         "TITLE", False,
     "Name of the series if this title belongs to one.",
     "The Adventure Series", 25),
    ("Series Number",       "TITLE", False,
     "Volume/book number within the series. Numeric only.",
     "3", 14),

    # ── PUBLISHER ────────────────────────────────────────────────────────
    ("Publisher",           "PUBLISHER", True,
     "Full legal publisher name exactly as registered.",
     "Compendium Inc.", 25),
    ("Imprint",             "PUBLISHER", False,
     "Imprint name if different from publisher. Otherwise repeat publisher name.",
     "Compendium", 20),
    ("Copyright Year",      "PUBLISHER", False,
     "4-digit year only.",
     "2026", 14),
    ("Copyright Owner",     "PUBLISHER", False,
     "Name of copyright holder if different from publisher.",
     "Compendium Inc.", 20),
    ("Copyright Statement", "PUBLISHER", False,
     "Full copyright statement if needed. Example: 2026 Compendium Inc.",
     "2026 Compendium Inc.", 30),
    # Country of Publication removed — defaults to US

    # ── FORMAT ───────────────────────────────────────────────────────────
    ("Product Format Code", "FORMAT", True,
     "ONIX code for physical format. See Code Lists tab. Common: BB=Hardback, BC=Paperback, EA=Digital download.",
     "BB", 20),
    ("Product Form Detail", "FORMAT", False,
     "ONIX product form detail code. See Code Lists tab. Example: B502=Trade paperback.",
     "B502", 18),
    ("Edition Number",      "FORMAT", False,
     "Numeric edition number. First editions: leave blank or enter 1.",
     "2", 14),
    ("Edition Type Code",   "FORMAT", False,
     "ONIX edition type. Example: REV=Revised, NED=New edition, ILL=Illustrated.",
     "REV", 16),

    # ── CONTRIBUTORS ─────────────────────────────────────────────────────
    ("Contrib Name 1",      "CONTRIBUTORS", True,
     "Format: Last, First (e.g. Smith, John). At least one contributor required.",
     "Clark, M.H.", 22),
    ("Contrib Role 1",      "CONTRIBUTORS", True,
     "ONIX role code. A01=Author, A12=Illustrator, B01=Editor, A38=Photographer. See Code Lists.",
     "A01", 14),
    ("Contrib Bio 1",       "CONTRIBUTORS", False,
     "Short author bio in plain text. 100-300 words recommended.",
     "M.H. Clark is a poet and author...", 40),
    ("Contrib Name 2",      "CONTRIBUTORS", False,
     "Format: Last, First",
     "Metzger, Cecile", 22),
    ("Contrib Role 2",      "CONTRIBUTORS", False,
     "ONIX role code. See Code Lists tab.",
     "A12", 14),
    ("Contrib Bio 2",       "CONTRIBUTORS", False,
     "Plain text bio.",
     "", 40),
    ("Contrib Name 3",      "CONTRIBUTORS", False,
     "Format: Last, First",
     "", 22),
    ("Contrib Role 3",      "CONTRIBUTORS", False,
     "ONIX role code.",
     "", 14),
    ("Contrib Bio 3",       "CONTRIBUTORS", False,
     "Plain text bio.",
     "", 40),
    ("Contrib Name 4",      "CONTRIBUTORS", False,
     "Format: Last, First",
     "", 22),
    ("Contrib Role 4",      "CONTRIBUTORS", False,
     "ONIX role code.",
     "", 14),
    ("Contrib Bio 4",       "CONTRIBUTORS", False,
     "Plain text bio.",
     "", 40),
    ("Contrib Name 5",      "CONTRIBUTORS", False,
     "Format: Last, First",
     "", 22),
    ("Contrib Role 5",      "CONTRIBUTORS", False,
     "ONIX role code.",
     "", 14),
    ("Contrib Bio 5",       "CONTRIBUTORS", False,
     "Plain text bio.",
     "", 40),

    # ── SUBJECT / CLASSIFICATION ──────────────────────────────────────────
    ("BISAC Code 1",        "SUBJECT", True,
     "Primary BISAC subject code. Format: XXX000000 (9 characters, no spaces). See bisac.com for full list. Maps to Amazon Browse Nodes.",
     "FAM021000", 16),
    ("BISAC Code 2",        "SUBJECT", False,
     "Secondary BISAC code.",
     "SEL016000", 16),
    ("BISAC Code 3",        "SUBJECT", False,
     "Additional BISAC code.",
     "SEL021000", 16),
    ("BISAC Code 4",        "SUBJECT", False,
     "Additional BISAC code.",
     "", 16),
    ("BISAC Code 5",        "SUBJECT", False,
     "Additional BISAC code.",
     "", 16),
    ("Subject Keywords",    "SUBJECT", False,
     "Semicolon-separated keywords. Used by Amazon and others for search indexing. No limit but keep focused.",
     "friendship; gift book; inspiring; holiday", 40),
    ("THEMA Code",          "SUBJECT", False,
     "THEMA subject code. Used by international channels (Waterstones, Gazelle, Edelweiss). See themabookcodes.info.",
     "D", 14),
    ("THEMA Description",   "SUBJECT", False,
     "Human-readable description of the THEMA code.",
     "Biography, Literature", 25),
    # THEMA Version removed — defaults to 1.6
    ("Audience Code",       "SUBJECT", False,
     "ONIX audience code. 01=General/Trade Adult, 02=Children, 03=Young Adult, 04=Primary/Elementary, 05=Secondary/High School. See Code Lists.",
     "01", 14),
    ("Audience Age From",   "SUBJECT", False,
     "Minimum age in years. Whole numbers only. Required for children's titles.",
     "4", 16),
    ("Audience Age To",     "SUBJECT", False,
     "Maximum age in years. Whole numbers only.",
     "8", 14),
    ("Grade Range",         "SUBJECT", False,
     "Grade range. Format: K-3, PK-2, 1-5, etc.",
     "PK-3", 14),

    # ── DESCRIPTIONS ─────────────────────────────────────────────────────
    ("Catalog Description", "DESCRIPTIONS", True,
     "Full marketing description / back cover copy. Plain text preferred; basic HTML (<p><b><i>) accepted. 200-500 words ideal. Amazon displays this on the product page.",
     "From the award-winning author...", 50),
    ("Short Description",   "DESCRIPTIONS", False,
     "Abbreviated description for search results and catalogs. 50-150 words. Plain text only.",
     "A touching story of friendship...", 40),
    ("Book Excerpt",        "DESCRIPTIONS", False,
     "Short excerpt from the book interior. Optional.",
     "", 40),
    ("TOC",                 "DESCRIPTIONS", False,
     "Table of contents. Semicolon-separated chapter titles or plain text.",
     "", 40),
    ("Promo Quote 1",       "DESCRIPTIONS", False,
     "Review quote or endorsement. Include attribution: 'Text here.' — Source Name",
     '"A beautiful book." — Publisher Weekly', 40),
    ("Promo Quote 2",       "DESCRIPTIONS", False,
     "Second review quote with attribution.",
     "", 40),
    ("Promo Quote 3",       "DESCRIPTIONS", False,
     "Third review quote with attribution.",
     "", 40),

    # ── DATES ────────────────────────────────────────────────────────────
    ("Pub Date",            "DATES", True,
     "Publication date. Format: YYYYMMDD (e.g. 20260115). This is the official publication date.",
     "20260115", 14),
    ("On Sale Date",        "DATES", False,
     "Date available for sale. Format: YYYYMMDD. Used for pre-orders and embargoes. If same as Pub Date, repeat it.",
     "20260115", 14),
    ("Ship Date",           "DATES", False,
     "Date copies ship to retailers. Format: YYYYMMDD. Usually 1-4 weeks before Pub Date.",
     "20251230", 14),

    # ── PHYSICAL ─────────────────────────────────────────────────────────
    ("Height (Inches)",     "PHYSICAL", False,
     "Book height in decimal inches. Include dust jacket if hardcover. Example: 9 or 9.5",
     "9", 14),
    ("Width (Inches)",      "PHYSICAL", False,
     "Book width in decimal inches.",
     "6.5", 14),
    ("Spine Thickness (Inches)","PHYSICAL", False,
     "Spine thickness in decimal inches.",
     "1.0", 18),
    ("Weight (Ounces)",     "PHYSICAL", False,
     "Book weight in decimal ounces.",
     "8.8", 14),
    ("Number of Pages",     "PHYSICAL", False,
     "Total page count including front and back matter.",
     "220", 14),
    ("Carton Quantity",     "PHYSICAL", False,
     "Number of copies per carton.",
     "24", 14),
    # Number of Pieces removed — defaults to 1
    ("Illustration Type",   "PHYSICAL", False,
     "ONIX illustration type code if applicable. See Code Lists.",
     "", 16),
    ("Illustration Notes",  "PHYSICAL", False,
     "Free-text description of illustrations.",
     "Full color throughout", 25),
    ("Number of Illustrations","PHYSICAL", False,
     "Total number of illustrations.",
     "32", 18),

    # ── PRICING ──────────────────────────────────────────────────────────
    ("US Price (USD)",      "PRICING", True,
     "US retail price in US dollars. Numeric only, no $ sign. Example: 17.99",
     "17.99", 16),
    ("CAD Price",           "PRICING", False,
     "Canadian retail price in CAD. Required for Amazon CA, Chapters-Indigo, TBM BookManager, ULS.",
     "24.99", 14),
    ("Intl Currency Code 1","PRICING", False,
     "ISO 4217 currency code for international price 1. Example: GBP, EUR, AUD.",
     "GBP", 18),
    ("Intl Price 1",        "PRICING", False,
     "Price in the currency specified in Intl Currency Code 1. Numeric only.",
     "14.99", 14),
    ("Intl Currency Code 2","PRICING", False,
     "ISO 4217 currency code for international price 2.",
     "EUR", 18),
    ("Intl Price 2",        "PRICING", False,
     "Price in the currency specified in Intl Currency Code 2. Numeric only.",
     "16.99", 14),
    # Discount Code removed — defaults to A

    # ── AVAILABILITY & STATUS ─────────────────────────────────────────────
    ("Publishing Status",   "AVAILABILITY", True,
     "ONIX publishing status code. 02=Forthcoming, 04=Active/Published, 06=Out of print, 07=Active no longer stocked. See Code Lists.",
     "04", 18),
    ("Product Availability","AVAILABILITY", True,
     "ONIX availability code. 20=Available, 21=In stock, 22=To order, 31=Out of stock indefinitely, 40=Not available. Must be consistent with Publishing Status.",
     "20", 20),
    # Notification Type removed — defaults to 03
    # Returns Code removed — defaults to Y
    ("Last Date of Return", "AVAILABILITY", False,
     "Last date returns accepted. Format: YYYYMMDD.",
     "", 18),

    # ── RIGHTS ───────────────────────────────────────────────────────────
    ("Sales Rights Type 1", "RIGHTS", False,
     "ONIX sales rights type. 01=Exclusive, 02=Non-exclusive, 03=Not for sale. See Code Lists.",
     "01", 18),
    ("Rights Territory 1",  "RIGHTS", False,
     "Territory for Sales Rights Type 1. Use ISO country codes space-separated, or WORLD for worldwide.",
     "US CA", 18),
    ("Sales Rights Type 2", "RIGHTS", False,
     "Second sales rights type if applicable.",
     "", 18),
    ("Rights Territory 2",  "RIGHTS", False,
     "Territory for Sales Rights Type 2.",
     "", 18),

    # ── LANGUAGE ─────────────────────────────────────────────────────────
    ("Language Code",       "LANGUAGE", True,
     "ISO 639-2/B language code. English=eng, French=fre, Spanish=spa. Pre-filled as 'eng' — only change for non-English titles.",
     "eng", 14),

    # ── COVER IMAGE ──────────────────────────────────────────────────────
    ("Cover Image URL",     "COVER", False,
     "Optional. If cover is hosted at a public HTTPS URL, enter it here. Otherwise leave blank -- upload the image file named ISBN13.jpg (or .png) and the system will auto-match it.",
     "https://cdn.example.com/covers/9780999496008.jpg", 45),

    # ── DISTRIBUTION ─────────────────────────────────────────────────────
    # Supplier Role removed — defaults to "Exclusive Distributor to resellers and end-customers"
    # Barcode Type removed — defaults to "GTIN-13 - On back"
    # Choking Hazard removed — defaults to "No choking hazard warning necessary"

    # ── RELATED PRODUCTS ─────────────────────────────────────────────────
    ("Replaces ISBN",       "RELATED", False,
     "ISBN13 of the title this product replaces (e.g. previous edition).",
     "", 16),
    ("Replaced by ISBN",    "RELATED", False,
     "ISBN13 of the title that replaces this product.",
     "", 16),
]

SECTIONS = [
    "IDENTIFICATION", "TITLE", "PUBLISHER", "FORMAT", "CONTRIBUTORS",
    "SUBJECT", "DESCRIPTIONS", "DATES", "PHYSICAL", "PRICING",
    "AVAILABILITY", "RIGHTS", "LANGUAGE", "COVER", "DISTRIBUTION", "RELATED"
]

SECTION_COLORS = {
    "IDENTIFICATION": "1F4E79",
    "TITLE":          "2E5FA3",
    "PUBLISHER":      "375623",
    "FORMAT":         "843C0C",
    "CONTRIBUTORS":   "7030A0",
    "SUBJECT":        "C00000",
    "DESCRIPTIONS":   "006064",
    "DATES":          "BF8F00",
    "PHYSICAL":       "1F4E79",
    "PRICING":        "375623",
    "AVAILABILITY":   "843C0C",
    "RIGHTS":         "7030A0",
    "LANGUAGE":       "2E5FA3",
    "COVER":          "C00000",
    "DISTRIBUTION":   "006064",
    "RELATED":        "BF8F00",
}

# ── Row 1: Section labels ─────────────────────────────────────────────────
# ── Row 2: Column headers ─────────────────────────────────────────────────
# ── Row 3: Format / instruction row ──────────────────────────────────────
# ── Row 4+: Data ─────────────────────────────────────────────────────────

current_section = None
section_start_col = 1

section_row   = ws.row_dimensions[1]
header_row    = ws.row_dimensions[2]
format_row    = ws.row_dimensions[3]
section_row.height = 18
header_row.height  = 30
format_row.height  = 52

for col_idx, (hdr, section, required, fmt_note, example, width) in enumerate(COLUMNS, start=1):
    col_letter = get_column_letter(col_idx)

    # ── Section header (row 1) ────────────────────────────────────────────
    sec_cell = ws.cell(row=1, column=col_idx)
    sec_color = SECTION_COLORS.get(section, NAVY)
    sec_cell.fill = fill(sec_color)
    sec_cell.font = Font(name="Arial", bold=True, color="FFFFFF", size=9)
    sec_cell.alignment = Alignment(horizontal="center", vertical="center")
    sec_cell.border = thin_border()
    # Only write section name in first column of each section
    if section != current_section:
        sec_cell.value = section
        current_section = section
    else:
        sec_cell.value = ""

    # ── Column header (row 2) ─────────────────────────────────────────────
    hdr_cell = ws.cell(row=2, column=col_idx)
    if required:
        hdr_cell.value = f"* {hdr}"
        hdr_cell.font = Font(name="Arial", bold=True, color="FFFFFF", size=9)
        hdr_cell.fill = fill("C00000")  # Red for required
    else:
        hdr_cell.value = hdr
        hdr_cell.font = Font(name="Arial", bold=True, color="FFFFFF", size=9)
        hdr_cell.fill = fill(DARK_BLUE)
    hdr_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    hdr_cell.border = thin_border()

    # ── Format note (row 3) ───────────────────────────────────────────────
    fmt_cell = ws.cell(row=3, column=col_idx)
    fmt_cell.value = fmt_note
    fmt_cell.font = Font(name="Arial", size=8, italic=True, color="404040")
    fmt_cell.fill = fill(PALE_BLUE)
    fmt_cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    fmt_cell.border = thin_border()

    # ── Column width ──────────────────────────────────────────────────────
    ws.column_dimensions[col_letter].width = width

# ── Data rows (rows 4-53 = 50 blank rows pre-formatted) ──────────────────
# Find Language Code column index for pre-fill
lang_col_idx = next(
    (i for i, (h, *_) in enumerate(COLUMNS, start=1) if h == "Language Code"),
    None
)

for row_idx in range(4, 54):
    for col_idx, (hdr, section, required, fmt_note, example, width) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=row_idx, column=col_idx)
        cell.font = cell_font(size=9)
        cell.alignment = Alignment(vertical="top", wrap_text=False)
        cell.border = thin_border()
        if row_idx % 2 == 0:
            cell.fill = fill("F7FBFF")
        else:
            cell.fill = fill(WHITE)
        # Pre-fill Language Code with "eng"
        if col_idx == lang_col_idx:
            cell.value = "eng"
            cell.font = Font(name="Arial", size=9, color="808080", italic=True)

# ── Legend row above section headers ────────────────────────────────────
# Insert a legend at top
ws.insert_rows(1)

legend_texts = [
    ("* Red header = REQUIRED field", "C00000"),
    ("Blue header = Optional field",  DARK_BLUE),
    ("Row 3 = Formatting instructions", "404040"),
    ("Dates format: YYYYMMDD",        "404040"),
    ("Do NOT add $ signs to prices",  "404040"),
]

ws.cell(row=1, column=1).value = "APG DISTRIBUTION — MASTER METADATA TEMPLATE   |   * = REQUIRED   |   Dates: YYYYMMDD   |   Prices: numeric only, no $ sign   |   See 'Instructions' and 'Code Lists' tabs"
ws.cell(row=1, column=1).font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
ws.cell(row=1, column=1).fill = fill(NAVY)
ws.cell(row=1, column=1).alignment = Alignment(horizontal="left", vertical="center")
ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(COLUMNS))
ws.row_dimensions[1].height = 22

# Now rows are: 1=legend, 2=section, 3=header, 4=format, 5+=data
# ════════════════════════════════════════════════════════════════════════════
# SHEET 2 — INSTRUCTIONS
# ════════════════════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("Instructions")
ws2.column_dimensions["A"].width = 25
ws2.column_dimensions["B"].width = 12
ws2.column_dimensions["C"].width = 12
ws2.column_dimensions["D"].width = 55
ws2.column_dimensions["E"].width = 30
ws2.column_dimensions["F"].width = 20

# Title
title_cell = ws2.cell(row=1, column=1, value="APG DISTRIBUTION — METADATA TEMPLATE FIELD REFERENCE")
title_cell.font = Font(name="Arial", bold=True, color="FFFFFF", size=12)
title_cell.fill = fill(NAVY)
title_cell.alignment = Alignment(horizontal="left", vertical="center")
ws2.merge_cells("A1:F1")
ws2.row_dimensions[1].height = 24

# Subtitle
ws2.cell(row=2, column=1).value = "Fields marked * are REQUIRED. All other fields are optional but improve channel discoverability."
ws2.cell(row=2, column=1).font = Font(name="Arial", italic=True, color="404040", size=9)
ws2.merge_cells("A2:F2")
ws2.row_dimensions[2].height = 16

# Header row
headers = ["Field Name", "Section", "Required?", "Instructions", "Example Value", "ONIX Element"]
header_row = 3
for col_idx, hdr in enumerate(headers, start=1):
    c = ws2.cell(row=header_row, column=col_idx)
    c.value = hdr
    c.font = Font(name="Arial", bold=True, color="FFFFFF", size=9)
    c.fill = fill(DARK_BLUE)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border = thin_border()
ws2.row_dimensions[header_row].height = 18

ONIX_ELEMENTS = {
    "ISBN10":              "ProductIdentifier / IDValue (type 02)",
    "ISBN13":              "ProductIdentifier / IDValue (type 15) — RecordReference",
    "GTIN":                "ProductIdentifier / IDValue (type 03)",
    "UPC":                 "ProductIdentifier / IDValue (type 04)",
    "Work ID":             "ProductIdentifier / IDValue (type 01)",
    "eISBN10 for Print":   "RelatedProduct / ProductIdentifier",
    "eISBN13 for Print":   "RelatedProduct / ProductIdentifier",
    "Title":               "TitleDetail / TitleText (TitleType 01)",
    "Subtitle":            "TitleDetail / Subtitle",
    "Series Name":         "Collection / TitleDetail / TitleText",
    "Series Number":       "Collection / CollectionSequence / CollectionSequenceValue",
    "Publisher":           "Publisher / PublisherName (role 01)",
    "Imprint":             "Imprint / ImprintName",
    "Copyright Year":      "CopyrightStatement / CopyrightYear",
    "Copyright Owner":     "CopyrightStatement / CopyrightOwner",
    "Copyright Statement": "CopyrightStatement / CopyrightStatement",
    "Country of Publication": "CountryOfPublication",
    "Product Format Code": "ProductForm",
    "Product Form Detail": "ProductFormDetail",
    "Edition Number":      "EditionNumber",
    "Edition Type Code":   "EditionType",
    "Contrib Name 1":      "Contributor / PersonName",
    "Contrib Role 1":      "Contributor / ContributorRole",
    "Contrib Bio 1":       "Contributor / BiographicalNote",
    "BISAC Code 1":        "Subject / SubjectCode (scheme 10 = BISAC)",
    "Subject Keywords":    "Subject / SubjectHeadingText (scheme 20)",
    "THEMA Code":          "Subject / SubjectCode (scheme 93 = Thema)",
    "Audience Code":       "Audience / AudienceCodeValue (type 01)",
    "Audience Age From":   "AudienceRange / AudienceRangeValue (precision 03)",
    "Audience Age To":     "AudienceRange / AudienceRangeValue (precision 04)",
    "Grade Range":         "AudienceRange (SchemeIdentifier 11)",
    "Catalog Description": "TextContent / Text (TextType 03)",
    "Short Description":   "TextContent / Text (TextType 02)",
    "Book Excerpt":        "TextContent / Text (TextType 24)",
    "TOC":                 "TextContent / Text (TextType 04)",
    "Promo Quote 1":       "TextContent / Text (TextType 06)",
    "Pub Date":            "PublishingDate / Date (role 01)",
    "On Sale Date":        "PublishingDate / Date (role 02)",
    "Ship Date":           "PublishingDate / Date (role 08)",
    "Height (Inches)":     "Measure / Measurement (type 01)",
    "Width (Inches)":      "Measure / Measurement (type 02)",
    "Spine Thickness (Inches)": "Measure / Measurement (type 03)",
    "Weight (Ounces)":     "Measure / Measurement (type 08)",
    "Number of Pages":     "Extent / ExtentValue (type 00)",
    "Carton Quantity":     "Pack / PackQuantity",
    "Number of Pieces":    "ProductComposition",
    "US Price (USD)":      "Price / PriceAmount (CurrencyCode USD)",
    "CAD Price":           "Price / PriceAmount (CurrencyCode CAD)",
    "Intl Currency Code 1":"Price / CurrencyCode",
    "Intl Price 1":        "Price / PriceAmount",
    "Discount Code":       "DiscountCoded / DiscountCode",
    "Publishing Status":   "PublishingStatus",
    "Product Availability":"ProductAvailability",
    "Notification Type":   "NotificationType",
    "Returns Code":        "ReturnsConditions / ReturnsCode",
    "Sales Rights Type 1": "SalesRights / SalesRightsType",
    "Rights Territory 1":  "SalesRights / Territory / CountriesIncluded",
    "Language Code":       "Language / LanguageCode (role 01)",
    "Cover Image URL":     "SupportingResource / ResourceLink (ContentType 01)",
    "Supplier Role":       "Supplier / SupplierRole",
}

data_row = 4
for col_idx, (hdr, section, required, fmt_note, example, width) in enumerate(COLUMNS, start=1):
    r = data_row
    c1 = ws2.cell(row=r, column=1, value=hdr)
    c2 = ws2.cell(row=r, column=2, value=section)
    c3 = ws2.cell(row=r, column=3, value="YES" if required else "Optional")
    c4 = ws2.cell(row=r, column=4, value=fmt_note)
    c5 = ws2.cell(row=r, column=5, value=example)
    c6 = ws2.cell(row=r, column=6, value=ONIX_ELEMENTS.get(hdr, ""))

    for cell in [c1, c2, c3, c4, c5, c6]:
        cell.font = Font(name="Arial", size=9)
        cell.alignment = Alignment(vertical="top", wrap_text=True)
        cell.border = thin_border()
        if r % 2 == 0:
            cell.fill = fill(GREY_HDR)

    c1.font = Font(name="Arial", size=9, bold=True)
    if required:
        c3.font = Font(name="Arial", size=9, bold=True, color="C00000")
        c3.fill = fill("FFF0F0")
    else:
        c3.font = Font(name="Arial", size=9, color="375623")

    ws2.row_dimensions[r].height = 42
    data_row += 1

ws2.freeze_panes = "A4"

# ════════════════════════════════════════════════════════════════════════════
# SHEET 3 — CODE LISTS
# ════════════════════════════════════════════════════════════════════════════
ws3 = wb.create_sheet("Code Lists")
ws3.column_dimensions["A"].width = 22
ws3.column_dimensions["B"].width = 14
ws3.column_dimensions["C"].width = 40
ws3.column_dimensions["D"].width = 5
ws3.column_dimensions["E"].width = 22
ws3.column_dimensions["F"].width = 14
ws3.column_dimensions["G"].width = 40

title3 = ws3.cell(row=1, column=1, value="APG DISTRIBUTION — ONIX CODE REFERENCE")
title3.font = Font(name="Arial", bold=True, color="FFFFFF", size=12)
title3.fill = fill(NAVY)
title3.alignment = Alignment(horizontal="left", vertical="center")
ws3.merge_cells("A1:G1")
ws3.row_dimensions[1].height = 24

code_sections = [
    ("PRODUCT FORMAT CODE (ProductForm)", [
        ("BB", "Hardback / Hardcover"),
        ("BC", "Paperback / Softcover"),
        ("BA", "Book — format unspecified"),
        ("BG", "Spiral bound"),
        ("BH", "Pamphlet / Saddle-stitched"),
        ("EA", "Digital download (eBook)"),
        ("AJ", "Audiobook MP3-CD"),
        ("AC", "Audiobook CD"),
    ]),
    ("PRODUCT FORM DETAIL (ProductFormDetail)", [
        ("B102", "Large format paperback"),
        ("B104", "Mass market paperback"),
        ("B201", "Coloring/activity book"),
        ("B401", "Library bound"),
        ("B402", "Reinforced binding"),
        ("B502", "Trade paperback (US)"),
        ("B504", "Digest format"),
        ("B601", "Board book"),
        ("B602", "Picture book"),
        ("B604", "Pop-up book"),
        ("B608", "Novelty book"),
        ("B610", "Cloth book"),
        ("B611", "Bath book"),
    ]),
    ("PUBLISHING STATUS", [
        ("02", "Forthcoming / Pre-publication"),
        ("03", "Postponed indefinitely"),
        ("04", "Active / Published"),
        ("05", "No longer our product"),
        ("06", "Out of print"),
        ("07", "Active — publisher no longer stocks"),
        ("08", "Inactive — awaiting reprint"),
        ("09", "Active — publisher direct only"),
        ("10", "Remaindered"),
        ("11", "Withdrawn from sale"),
        ("12", "Not yet published — confirmed pub date"),
    ]),
    ("PRODUCT AVAILABILITY", [
        ("20", "Available / In stock"),
        ("21", "In stock — ships within 2 days"),
        ("22", "To order / Available 3-30 days"),
        ("23", "Manufactured on demand (POD)"),
        ("31", "Out of stock indefinitely"),
        ("32", "Out of stock — no info on availability"),
        ("33", "Temporarily unavailable"),
        ("40", "Not available — reason unspecified"),
        ("41", "Not available in this market"),
        ("42", "Not available — publisher address unknown"),
    ]),
    ("CONTRIBUTOR ROLE CODES", [
        ("A01", "Author / By"),
        ("A02", "With / Co-author"),
        ("A03", "Screenwriter"),
        ("A06", "Composer"),
        ("A12", "Illustrator"),
        ("A14", "Photographer"),
        ("A24", "Preface by"),
        ("A25", "Prologue by"),
        ("A27", "Foreword by"),
        ("A38", "Original author (translation)"),
        ("B01", "Editor"),
        ("B06", "Translator"),
        ("B09", "Series editor"),
        ("B10", "Edited and translated by"),
        ("C01", "Production editor"),
        ("D01", "Producer"),
        ("E07", "Read by (audiobook narrator)"),
    ]),
    ("AUDIENCE CODE", [
        ("01", "General / Trade adult"),
        ("02", "Children / Juvenile"),
        ("03", "Young adult (YA)"),
        ("04", "Primary / Elementary school"),
        ("05", "Secondary / High school"),
        ("06", "College / Higher education"),
        ("07", "Professional / Academic"),
        ("08", "English language learners (ELT/ESL)"),
    ]),
    ("NOTIFICATION TYPE", [
        ("01", "Early notification — advance info, may change"),
        ("02", "Advance notification — confirmed detail"),
        ("03", "Notification / Confirmation — at or after publication"),
        ("04", "Update — partial update to existing record"),
        ("05", "Delete / Cancel record"),
    ]),
    ("SALES RIGHTS TYPE", [
        ("01", "Exclusive rights in territory"),
        ("02", "Non-exclusive rights in territory"),
        ("03", "Not for sale in territory"),
        ("04", "Exclusive rights worldwide"),
        ("05", "Non-exclusive rights worldwide"),
        ("06", "Not for sale worldwide"),
    ]),
    ("LANGUAGE CODE (ISO 639-2/B)", [
        ("eng", "English"),
        ("fre", "French"),
        ("spa", "Spanish"),
        ("ger", "German"),
        ("ita", "Italian"),
        ("por", "Portuguese"),
        ("chi", "Chinese (simplified/traditional — use zho for ONIX 3)"),
        ("jpn", "Japanese"),
        ("kor", "Korean"),
        ("ara", "Arabic"),
        ("rus", "Russian"),
    ]),
    ("RETURNS CODE", [
        ("Y", "Returnable"),
        ("N", "Non-returnable / Final sale"),
        ("C", "Conditional — contact supplier"),
    ]),
    ("COMMON ISO COUNTRY CODES", [
        ("US", "United States"),
        ("CA", "Canada"),
        ("GB", "United Kingdom"),
        ("AU", "Australia"),
        ("NZ", "New Zealand"),
        ("IE", "Ireland"),
        ("ZA", "South Africa"),
        ("IN", "India"),
        ("WORLD", "Worldwide (use in SalesRights territory)"),
    ]),
    ("PRICE TYPE", [
        ("01", "Retail price (RRP incl. tax) — use for US/CA consumer"),
        ("02", "Retail price exc. tax"),
        ("41", "Publisher price exc. tax (net)"),
    ]),
    ("COVER IMAGE REQUIREMENTS", [
        ("Min dimension", "2,500px height × 1,600px width recommended"),
        ("Resolution", "300 DPI minimum"),
        ("Format", "JPEG (RGB color only — no CMYK)"),
        ("Max file size", "5 MB"),
        ("Filename convention", "ISBN13.jpg or ISBN13.png — e.g. 9780999496008.jpg"),
        ("Background", "Covers with white backgrounds: add 3-4px gray border"),
        ("ONIX element", "SupportingResource / ResourceLink (ContentType 01, ResourceMode 03)"),
    ]),
]

# Two-column layout for code tables
current_row = 3
col_a, col_b, col_c = 1, 2, 3
col_e, col_f, col_g = 5, 6, 7

left_sections  = code_sections[0::2]   # even-indexed
right_sections = code_sections[1::2]   # odd-indexed
max_sections   = max(len(left_sections), len(right_sections))

current_row = 3

def write_code_section(ws3, start_row, sections, col_label, col_code, col_desc):
    r = start_row
    for section_title, codes in sections:
        # Section header
        sec_hdr = ws3.cell(row=r, column=col_label, value=section_title)
        sec_hdr.font = Font(name="Arial", bold=True, color="FFFFFF", size=9)
        sec_hdr.fill = fill(DARK_BLUE)
        sec_hdr.alignment = Alignment(horizontal="left", vertical="center")
        sec_hdr.border = thin_border()
        ws3.cell(row=r, column=col_code).fill = fill(DARK_BLUE)
        ws3.cell(row=r, column=col_code).border = thin_border()
        ws3.cell(row=r, column=col_desc).fill = fill(DARK_BLUE)
        ws3.cell(row=r, column=col_desc).border = thin_border()
        ws3.row_dimensions[r].height = 16
        r += 1

        # Column sub-headers
        for col, val in [(col_label, "Code"), (col_code, "Value"), (col_desc, "Description")]:
            c = ws3.cell(row=r, column=col, value=val)
            c.font = Font(name="Arial", bold=True, color="FFFFFF", size=8)
            c.fill = fill(MED_BLUE)
            c.alignment = Alignment(horizontal="center")
            c.border = thin_border()
        ws3.row_dimensions[r].height = 14
        r += 1

        for i, (code, desc) in enumerate(codes):
            c1 = ws3.cell(row=r, column=col_label, value=code)
            c2 = ws3.cell(row=r, column=col_code, value="")
            c3 = ws3.cell(row=r, column=col_desc, value=desc)
            for c in [c1, c2, c3]:
                c.font = Font(name="Arial", size=9)
                c.border = thin_border()
                c.alignment = Alignment(vertical="center")
                if i % 2 == 0:
                    c.fill = fill(GREY_HDR)
            ws3.row_dimensions[r].height = 14
            r += 1

        r += 1  # blank spacer
    return r

# Write left column (A-C) and right column (E-G) in parallel
left_pairs  = list(enumerate(code_sections[0::2]))
right_pairs = list(enumerate(code_sections[1::2]))

lr = 3
rr = 3
for i in range(max_sections):
    l_end = lr
    r_end = rr
    if i < len(left_pairs):
        _, sec = left_pairs[i]
        rows_needed = 2 + len(sec[1]) + 2
        write_code_section(ws3, lr, [sec], col_a, col_b, col_c)
        l_end = lr + rows_needed
    if i < len(right_pairs):
        _, sec = right_pairs[i]
        rows_needed = 2 + len(sec[1]) + 2
        write_code_section(ws3, rr, [sec], col_e, col_f, col_g)
        r_end = rr + rows_needed
    lr = l_end
    rr = r_end

ws3.freeze_panes = "A3"
ws3.column_dimensions["D"].width = 3   # spacer column

# ── Save ─────────────────────────────────────────────────────────────────
output_path = r"c:\Users\jflanagan\OneDrive - WFS, LLC\Desktop\EA_Claude\projects\metadata-exporter\templates\APG_Metadata_Template_v2.xlsx"
wb.save(output_path)
print(f"Saved: {output_path}")
